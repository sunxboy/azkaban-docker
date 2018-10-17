#-*- coding: utf-8 -*-
# 依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/
# author:huhao
# date:2017-12-27
#

from openpyxl import load_workbook

import csv
import os
import sys
import time
import xlrd
import codecs

def ExcelToCsv(excelfilepath):
    ## D:\xlspy
    filepath = os.path.dirname(excelfilepath)
    ## aaaaa.xlsx
    filefullname = os.path.basename(excelfilepath)
    filename = os.path.splitext(filefullname)[0]
    subfix = os.path.splitext(filefullname)[1]

    print("PythonExcel2Csv>>>> Input:" + unicode(excelfilepath, "GBK"))

    if subfix == ".xlsx":
        return Excel2007ToCsv(excelfilepath, filepath, filename)
    elif subfix == ".xls":
        return Excel2003ToCsv(excelfilepath, filepath, filename)
    else:
        print("PythonExcel2Csv>>>> Not supported file format!")
        return None

def Excel2003ToCsv(excelfilepath, filepath, filename):
    starttime = time.time()
    csvfile = ""
    try:
        workbook = xlrd.open_workbook(excelfilepath)
        firstsheet = workbook.sheet_by_index(0)
        print(workbook.sheet_names()[0])

        ## 创建目录 D:\xlspy\aaaaa\exceltocsv\{sheetname}.csv
        csvpath = filepath + "/" + filename + "/exceltocsv"
        csv_filename = 'sheet1.csv'
        ##csv_filename = '{sheet1}.csv'.format(
        ##    sheet1=workbook.sheet_names()[0].replace(' ', '_').encode('utf8'))

        if os.path.exists(csvpath) == False:
            os.makedirs(csvpath)

        csvfile = csvpath + "/" + csv_filename
        if os.path.exists(csvfile) == False:
            ##csv_file = file(csvfile, 'wb')
            csv_file = codecs.open(csvfile, 'wb', 'utf-16')
            csv_file_writer = csv.writer(csv_file)

            for row in xrange(0, firstsheet.nrows):
                rows = firstsheet.row_values(row)
                row_container = []

                def _tostr(cell):
                    if type(u'') == type(cell):
                        val = "%s" % cell.encode('utf8').replace('\n', '')
                        val =  val[:-2]  if val.endswith('.0') else val
                        if val == "None":
                            return ""
                        else:
                            return val
                    else:
                        val = "%s" % str(cell)
                        val = val[:-2] if val.endswith('.0') else val
                        if val == "None":
                            return ""
                        else:
                            return val

                row_container.append('\t'.join([_tostr(cell) for cell in rows]))
                csv_file_writer.writerow(row_container)
            csv_file.close()
        else:
            print("PythonExcel2Csv>>>> The csv file exist!!")
    except Exception as e:
        print(e)

    endtime = time.time() - starttime
    print("PythonExcel2Csv>>>> Spend time: " + str(endtime))
    return unicode(csvfile, "GBK")

def Excel2007ToCsv(excelfilepath, filepath, filename):
    starttime1 = time.time()
    csvfile1 = ""
    try:
        wb = load_workbook(filename=excelfilepath, read_only=True)
        ## 暂时只支持取第一个sheet, 以后再扩展
        ##firstsheet = xlsx_file_reader.get_sheet_names()[0]
        ##firstsheet = xlsx_file_reader.get_sheet_by_name("Sheet1")

        firstsheet1 = wb.get_sheet_by_name(wb.sheetnames[0])  # index为0为第一张表
        print(wb.sheetnames[0])

        ## 创建目录 D:\xlspy\aaaaa\exceltocsv\{sheetname}.csv
        csvpath1 = filepath + "/" + filename + "/exceltocsv"
        csv_filename1 = 'sheet1.csv'
        ##csv_filename1 = '{sheet111}.csv'.format(
        ##    sheet111=(wb.sheetnames[0]).replace(' ', '_'))

        if os.path.exists(csvpath1) == False:
            os.makedirs(csvpath1)

        csvfile1 = csvpath1 + "/" + csv_filename1
        if os.path.exists(csvfile1) == False:
            ##csv_file1 = file(csvfile1, 'wb')
            csv_file1 = codecs.open(csvfile1, 'wb', 'utf-16')
            csv_file_writer1 = csv.writer(csv_file1, delimiter='\t')

            for row1 in firstsheet1.rows:
                row_container1 = []
                for cell in row1:
                    if type(cell.value) == unicode:
                        val = "%s" % cell.value.encode('utf-8').replace('\n', '')
                        val = val[:-2] if val.endswith('.0') else val
                        if val == "None":
                            row_container1.append("");
                        else:
                            row_container1.append(val);
                    else:
                        val = "%s" % str(cell.value)
                        val = val[:-2] if val.endswith('.0') else val
                        if val == "None":
                            row_container1.append("");
                        else:
                            row_container1.append(val);
                csv_file_writer1.writerow(row_container1)
            csv_file1.close()
        else:
            print("PythonExcel2Csv>>>> The csv file exist!!")

    except Exception as e:
        print(e)

    endtime1 = time.time() - starttime1
    print("PythonExcel2Csv>>>> Spend time: " + str(endtime1))
    return unicode(csvfile1, "GBK")

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('PythonExcel2Csv>>>> usage: ExcelToCsv <xls or xlsx file path>')
    else:
        reload(sys)
        sys.setdefaultencoding('utf-8')
        print('PythonExcel2Csv>>>> success:' + ExcelToCsv(sys.argv[1]))
    sys.exit(0)